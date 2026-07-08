"""
Excel exporter for the AI Lead Insights pipeline.

Writes a .xlsx whose columns mirror the reference 'AI Lead Insights' sheet, so
the output drops straight into the client's existing workflow. Column order and
headers are fixed; values come from enriched record dicts (see insights_enrich).
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (header, record-key, column-width) in reference order.
# A record-key of None means the column is filled specially (see below).
COLUMNS = [
    ("Company Domain",                            "company_domain",   22),
    ("First Name",                                "first_name",       14),
    ("Last Name",                                 "last_name",        14),
    ("LinkedIn Profile URL",                      "linkedin_url",     34),
    ("Email Address",                             "email",            30),
    ("Company Size",                              "company_size",     14),
    ("Job Title",                                 "job_title",        24),
    ("Seniority Level",                           "seniority",        16),
    ("Industry",                                  "industry",         22),
    ("Location",                                  "location",         22),
    ("Company Name",                              "company_name",     22),
    ("Account and Company Topline Description",   "company_topline",  50),
    ("Latest Press Releases (Summary Sample)",    "company_press",    50),
    ("How to target the company",                 "how_to_target",    50),
    ("Interest / Topic Selected",                 "interest_topic",   26),
    ("Message From Lead / Form Input",            "message_from_lead",26),
    ("Position & Title",                          "position_bio",     50),
    ("Past 3 Work Experience",                    "past_experience",  50),
    ("Latest Press Releases (Summary)",           "person_press",     50),
    ("CVP against Lead's Profile",                "cvp",              50),
    ("Personalised Sales Paragraph",              "sales_paragraph",  60),
    ("Proposed Heat Rank",                        "heat_rank",        16),
    ("Sample Scoring Logic",                      None,               40),
]

# Verification columns — appended only when the records carry verification data
# (i.e. the run was not invoked with --no-verify). See verify.py.
VERIFICATION_COLUMNS = [
    ("Email Status",       "email_result",       14),  # deliverable/risky/undeliverable
    ("Data Confidence",    "data_confidence",    14),  # high/med/low (objective, not self-reported)
    ("Verification Notes", "verification_notes", 46),
]

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_WRAP_TOP    = Alignment(wrap_text=True, vertical="top")


def _write_sheet(ws, columns: list, records: list[dict], scoring_logic: str) -> None:
    """Render one worksheet: styled header row + data rows."""
    for col_idx, (header, _key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for r, rec in enumerate(records, start=2):
        for col_idx, (_header, key, _width) in enumerate(columns, start=1):
            if key is None:
                # 'Sample Scoring Logic' — the rubric belongs to the run, not the
                # row; write it once on the first data row to keep the sheet clean.
                value = scoring_logic if r == 2 else ""
            else:
                value = rec.get(key, "")
                if isinstance(value, (list, dict)):
                    value = str(value)
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.alignment = _WRAP_TOP

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def write_xlsx(
    records: list[dict],
    client: dict,
    filename: str | None = None,
    sheet_name: str = "AI Insights",
) -> str:
    """Write enriched records to a formatted .xlsx, return the filename.

    If the records were verified (carry `data_confidence`), three verification
    columns are appended and any rows flagged `needs_review` are also written to
    a separate 'Review' sheet so reps triage them before sending.
    """
    if filename is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ai_lead_insights_{client.get('client_name','client').lower()}_{stamp}.xlsx"

    # Base reference columns + client-specific account columns + (if present)
    # verification columns.
    columns = COLUMNS + list(client.get("account_columns", []))
    verified = any("data_confidence" in rec for rec in records)
    if verified:
        columns = columns + VERIFICATION_COLUMNS

    scoring_logic = client.get("scoring_logic", "")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _write_sheet(ws, columns, records, scoring_logic)

    # Review sheet — only the rows needing human attention before outreach.
    if verified:
        flagged = [rec for rec in records if rec.get("needs_review")]
        if flagged:
            review = wb.create_sheet("Review")
            _write_sheet(review, columns, flagged, scoring_logic)

    wb.save(filename)
    return filename
