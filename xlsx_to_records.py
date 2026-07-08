#!/usr/bin/env python3
"""
Rebuild the viewer's _insights_records.json from an AI Lead Insights .xlsx.

Maps each column header back to its record key (the inverse of excel_export's
COLUMNS / account_columns / VERIFICATION_COLUMNS), so an exported sheet can
repopulate the Streamlit viewer without re-running (and re-paying for) enrichment.

Usage:
    python xlsx_to_records.py <path.xlsx> [--client oangle] [--out _insights_records.json]
"""
import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

import client_config
import excel_export

HERE = Path(__file__).parent


def header_to_key(client_key: str) -> dict:
    """Build {column header -> record key} for the active column layout."""
    profile = client_config.get_client(client_key)
    columns = (excel_export.COLUMNS
               + list(profile.get("account_columns", []))
               + excel_export.VERIFICATION_COLUMNS)
    return {header: key for (header, key, _w) in columns if key}


def records_from_xlsx(path: str, client_key: str = "oangle",
                      sheet: str = "AI Insights") -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    h2k = header_to_key(client_key)
    records = []
    for row in rows[1:]:
        if all(v in (None, "") for v in row):
            continue
        rec = {}
        for header, value in zip(headers, row):
            key = h2k.get(header)
            if key:
                rec[key] = "" if value is None else value
        records.append(rec)
    return records


def main():
    ap = argparse.ArgumentParser(description="Rebuild _insights_records.json from an xlsx")
    ap.add_argument("xlsx")
    ap.add_argument("--client", default="oangle")
    ap.add_argument("--out", default=str(HERE / "_insights_records.json"))
    args = ap.parse_args()

    records = records_from_xlsx(args.xlsx, args.client)
    Path(args.out).write_text(json.dumps(records, indent=2, default=str))
    print(f"Wrote {len(records)} record(s) → {args.out}")


if __name__ == "__main__":
    main()
